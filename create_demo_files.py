"""
Column Auto-Detection Demo
Shows how the system automatically identifies email and other columns
"""
import pandas as pd
import openpyxl
from datetime import datetime, timedelta

def create_demo_files():
    """Create demo Excel files with different column names"""
    
    demos = [
        {
            'filename': 'demo_standard.xlsx',
            'columns': ['Candidate Email', 'Interview Date', 'Interview Time', 'Interview Description', 'Status'],
            'description': 'Standard column names'
        },
        {
            'filename': 'demo_alternate1.xlsx',
            'columns': ['Email ID', 'Date', 'Time', 'Details', 'Sent'],
            'description': 'Alternate names - Email ID, Details, Sent'
        },
        {
            'filename': 'demo_alternate2.xlsx',
            'columns': ['Recipient', 'Schedule Date', 'Hour', 'Interview Info', 'Status'],
            'description': 'Alternate names - Recipient, Schedule Date, Hour'
        },
        {
            'filename': 'demo_alternate3.xlsx',
            'columns': ['Candidate', 'When', 'Timing', 'Subject', 'State'],
            'description': 'Alternate names - Candidate, When, Timing, Subject'
        }
    ]
    
    print("=" * 70)
    print("  CREATING DEMO FILES WITH DIFFERENT COLUMN NAMES")
    print("=" * 70)
    print()
    
    for demo in demos:
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Add headers
        for col_num, header in enumerate(demo['columns'], 1):
            ws.cell(row=1, column=col_num, value=header)
        
        # Add sample data
        ws.cell(row=2, column=1, value="test@example.com")
        ws.cell(row=2, column=2, value=(datetime.now() + timedelta(days=3)).strftime("%Y-%m-%d"))
        ws.cell(row=2, column=3, value="10:00 AM")
        ws.cell(row=2, column=4, value="Technical Interview - Python")
        ws.cell(row=2, column=5, value="")
        
        wb.save(demo['filename'])
        
        print(f"✓ Created: {demo['filename']}")
        print(f"  {demo['description']}")
        print(f"  Columns: {', '.join(demo['columns'])}")
        print()
    
    print("=" * 70)
    print("  ✅ ALL DEMO FILES CREATED!")
    print("=" * 70)
    print()
    print("📋 AUTO-DETECTION PATTERNS:")
    print()
    print("  📧 Email Column - Detects:")
    print("     • 'email', 'mail', 'e-mail', 'candidate', 'recipient'")
    print()
    print("  📅 Date Column - Detects:")
    print("     • 'date', 'day', 'when', 'schedule'")
    print()
    print("  ⏰ Time Column - Detects:")
    print("     • 'time', 'hour', 'timing'")
    print()
    print("  📝 Description Column - Detects:")
    print("     • 'description', 'detail', 'info', 'note', 'subject', 'topic'")
    print()
    print("  ✅ Status Column - Detects:")
    print("     • 'status', 'sent', 'state'")
    print()
    print("=" * 70)
    print()
    print("🎯 USAGE:")
    print("  1. Upload any demo file to the web app")
    print("  2. Columns will be auto-detected")
    print("  3. You can manually change if needed")
    print()


if __name__ == "__main__":
    create_demo_files()
