"""
Streamlit Web Interface for Interview Notification Scheduler
Allows users to upload Excel files and send interview notifications via UI
"""
import streamlit as st
import pandas as pd
import openpyxl
from datetime import datetime
import os
from email_sender import OutlookEmailer
from logger import EmailLogger
import tempfile


# Page configuration
st.set_page_config(
    page_title="Interview Notification Scheduler",
    page_icon="📧",
    layout="wide"
)

# Custom CSS
st.markdown("""
    <style>
    .main-header {
        font-size: 2.5rem;
        color: #1f77b4;
        text-align: center;
        padding: 1rem 0;
    }
    .success-box {
        padding: 1rem;
        background-color: #d4edda;
        border-left: 5px solid #28a745;
        margin: 1rem 0;
    }
    .error-box {
        padding: 1rem;
        background-color: #f8d7da;
        border-left: 5px solid #dc3545;
        margin: 1rem 0;
    }
    .info-box {
        padding: 1rem;
        background-color: #d1ecf1;
        border-left: 5px solid #0c5460;
        margin: 1rem 0;
    }
    </style>
""", unsafe_allow_html=True)

# Initialize session state
if 'uploaded_file_path' not in st.session_state:
    st.session_state.uploaded_file_path = None
if 'email_results' not in st.session_state:
    st.session_state.email_results = []
if 'show_results' not in st.session_state:
    st.session_state.show_results = False


def load_excel_data(file_path):
    """Load and display Excel data"""
    try:
        df = pd.read_excel(file_path)
        return df
    except Exception as e:
        st.error(f"Error loading Excel file: {str(e)}")
        return None


def detect_columns(df):
    """Auto-detect column mappings based on common patterns"""
    column_mapping = {
        'email': None,
        'date': None,
        'time': None,
        'description': None,
        'status': None
    }
    
    # Email patterns
    email_patterns = ['email', 'mail', 'e-mail', 'candidate', 'recipient']
    for col in df.columns:
        if any(pattern in col.lower() for pattern in email_patterns):
            column_mapping['email'] = col
            break
    
    # Date patterns
    date_patterns = ['date', 'day', 'when', 'schedule']
    for col in df.columns:
        if any(pattern in col.lower() for pattern in date_patterns) and 'time' not in col.lower():
            column_mapping['date'] = col
            break
    
    # Time patterns
    time_patterns = ['time', 'hour', 'timing']
    for col in df.columns:
        if any(pattern in col.lower() for pattern in time_patterns):
            column_mapping['time'] = col
            break
    
    # Description patterns
    desc_patterns = ['description', 'detail', 'info', 'note', 'subject', 'topic']
    for col in df.columns:
        if any(pattern in col.lower() for pattern in desc_patterns):
            column_mapping['description'] = col
            break
    
    # Status patterns
    status_patterns = ['status', 'sent', 'state']
    for col in df.columns:
        if any(pattern in col.lower() for pattern in status_patterns):
            column_mapping['status'] = col
            break
    
    return column_mapping


def send_emails_with_mapping(file_path, column_mapping):
    """Send emails using uploaded file with custom column mapping"""
    results = {
        'sent': [],
        'failed': [],
        'skipped': []
    }
    
    # Initialize logger
    logger = EmailLogger("streamlit_email_notifications.log")
    logger.log_session_start()
    
    # Load Excel file
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Find column indices
        header_row = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
        
        email_col = header_row.get(column_mapping['email'])
        date_col = header_row.get(column_mapping['date'])
        time_col = header_row.get(column_mapping['time'])
        desc_col = header_row.get(column_mapping['description'])
        status_col = header_row.get(column_mapping['status']) if column_mapping['status'] else None
        
        # Initialize Outlook emailer
        emailer = OutlookEmailer()
        if not emailer.connect():
            wb.close()
            return None
        
        # Process rows
        for row_num in range(2, ws.max_row + 1):
            email = ws.cell(row=row_num, column=email_col).value
            date = ws.cell(row=row_num, column=date_col).value if date_col else ""
            time_val = ws.cell(row=row_num, column=time_col).value if time_col else ""
            description = ws.cell(row=row_num, column=desc_col).value if desc_col else ""
            status = ws.cell(row=row_num, column=status_col).value if status_col else None
            
            # Skip if already sent or no email
            if status == "Sent" or not email:
                if status == "Sent":
                    results['skipped'].append({'email': str(email)})
                continue
            
            # Validate required fields
            if not all([email, date, time_val, description]):
                results['failed'].append({
                    'email': str(email) if email else 'Unknown',
                    'error': 'Missing required data'
                })
                continue
            
            # Prepare interview data
            interview_data = {
                'email': str(email).strip(),
                'date': str(date).strip(),
                'time': str(time_val).strip(),
                'description': str(description).strip()
            }
            
            # Send email
            if emailer.send_interview_notification(interview_data):
                # Mark as sent
                if status_col:
                    ws.cell(row=row_num, column=status_col, value="Sent")
                    wb.save(file_path)
                
                logger.log_email_sent(interview_data['email'])
                results['sent'].append({
                    'email': interview_data['email'],
                    'date': interview_data['date'],
                    'time': interview_data['time']
                })
            else:
                logger.log_email_failed(interview_data['email'], "Failed to send email")
                results['failed'].append({
                    'email': interview_data['email'],
                    'error': 'Failed to send email'
                })
        
        wb.close()
        logger.log_session_end(len(results['sent']), len(results['failed']))
        
    except Exception as e:
        logger.log_email_failed("System", str(e))
        return None
    
    return results


# Header
st.markdown('<h1 class="main-header">📧 Interview Notification Scheduler</h1>', unsafe_allow_html=True)
st.markdown("---")

# Sidebar
with st.sidebar:
    st.header("📋 Instructions")
    st.markdown("""
    **Step 1:** Upload your Excel file
    
    **Step 2:** Map columns to fields
    
    **Step 3:** Preview the interviews
    
    **Step 4:** Click 'Send Notifications'
    
    ---
    
    **Required Data:**
    - Email address
    - Interview date
    - Interview time  
    - Interview details/description
    
    **Optional:**
    - Status column (for tracking)
    """)
    
    st.markdown("---")
    
    if st.button("📄 Download Sample Template"):
        st.info("Run `python create_template.py` in terminal to create a template file.")

# Main content
col1, col2 = st.columns([2, 1])

with col1:
    st.header("📤 Upload Interview Data")
    uploaded_file = st.file_uploader(
        "Choose an Excel file (.xlsx)",
        type=['xlsx'],
        help="Upload an Excel file containing interview details"
    )

with col2:
    st.header("📊 Quick Stats")
    if uploaded_file is not None:
        # Save uploaded file temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            tmp_file.write(uploaded_file.getvalue())
            temp_file_path = tmp_file.name
            st.session_state.uploaded_file_path = temp_file_path
        
        # Load and display stats
        df = load_excel_data(temp_file_path)
        
        if df is not None:
            total_rows = len(df)
            # Try to find status column
            status_col_name = None
            for col in df.columns:
                if 'status' in col.lower() or 'sent' in col.lower():
                    status_col_name = col
                    break
            
            sent_count = len(df[df[status_col_name] == 'Sent']) if status_col_name and status_col_name in df.columns else 0
            pending_count = total_rows - sent_count
            
            st.metric("Total Interviews", total_rows)
            st.metric("Already Sent", sent_count, delta=None)
            st.metric("Pending", pending_count, delta=None)

# Display Excel content and column mapping
if uploaded_file is not None and st.session_state.uploaded_file_path:
    st.markdown("---")
    
    df = load_excel_data(st.session_state.uploaded_file_path)
    
    if df is not None:
        # Auto-detect columns
        detected_mapping = detect_columns(df)
        
        st.header("🗺️ Column Mapping")
        st.info("✨ Auto-detected columns below. Change if needed:")
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            email_col = st.selectbox(
                "📧 Email Column",
                options=df.columns.tolist(),
                index=df.columns.tolist().index(detected_mapping['email']) if detected_mapping['email'] in df.columns else 0
            )
        
        with col2:
            date_col = st.selectbox(
                "📅 Date Column",
                options=df.columns.tolist(),
                index=df.columns.tolist().index(detected_mapping['date']) if detected_mapping['date'] in df.columns else 0
            )
        
        with col3:
            time_col = st.selectbox(
                "⏰ Time Column",
                options=df.columns.tolist(),
                index=df.columns.tolist().index(detected_mapping['time']) if detected_mapping['time'] in df.columns else 0
            )
        
        with col4:
            desc_col = st.selectbox(
                "📝 Description Column",
                options=df.columns.tolist(),
                index=df.columns.tolist().index(detected_mapping['description']) if detected_mapping['description'] in df.columns else 0
            )
        
        col5, col6 = st.columns([1, 3])
        with col5:
            status_options = ["None"] + df.columns.tolist()
            status_col = st.selectbox(
                "✅ Status Column (Optional)",
                options=status_options,
                index=status_options.index(detected_mapping['status']) if detected_mapping['status'] in status_options else 0
            )
        
        # Store mapping
        column_mapping = {
            'email': email_col,
            'date': date_col,
            'time': time_col,
            'description': desc_col,
            'status': status_col if status_col != "None" else None
        }
        
        st.markdown("---")
        st.header("📋 Interview Data Preview")
        
        # Display data with color coding
        def highlight_status(row):
            if status_col and status_col != "None" and status_col in row.index:
                if row[status_col] == 'Sent':
                    return ['background-color: #d4edda'] * len(row)
            return [''] * len(row)
        
        styled_df = df.style.apply(highlight_status, axis=1)
        st.dataframe(styled_df, use_container_width=True, height=400)
        
        # Filter pending interviews
        if status_col and status_col != "None":
            pending_df = df[df[status_col] != 'Sent']
        else:
            pending_df = df
        
        pending_df = pending_df[pending_df[email_col].notna()]
        
        if len(pending_df) > 0:
            st.markdown("---")
            st.header("🚀 Send Notifications")
            
            col1, col2, col3 = st.columns([2, 1, 1])
            
            with col1:
                st.info(f"📨 Ready to send {len(pending_df)} notification(s)")
            
            with col2:
                if st.button("✉️ Send All Notifications", type="primary", use_container_width=True):
                    with st.spinner("Sending emails..."):
                        results = send_emails_with_mapping(st.session_state.uploaded_file_path, column_mapping)
                        
                        if results is not None:
                            st.session_state.email_results = results
                            st.session_state.show_results = True
                            st.rerun()
                        else:
                            st.error("Failed to connect to Outlook or process file!")
            
            with col3:
                if st.button("🔄 Refresh Data", use_container_width=True):
                    st.rerun()
        else:
            st.markdown('<div class="info-box">ℹ️ No pending interviews to send. All notifications have been sent!</div>', unsafe_allow_html=True)

# Display results
if st.session_state.show_results:
    st.markdown("---")
    st.header("📊 Sending Results")
    
    results = st.session_state.email_results
    
    # Summary metrics
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric("✅ Sent Successfully", len(results['sent']), delta=None)
    
    with col2:
        st.metric("❌ Failed", len(results['failed']), delta=None)
    
    with col3:
        st.metric("⏭️ Skipped", len(results['skipped']), delta=None)
    
    # Detailed results
    if results['sent']:
        st.subheader("✅ Successfully Sent")
        sent_df = pd.DataFrame(results['sent'])
        st.dataframe(sent_df, use_container_width=True)
    
    if results['failed']:
        st.subheader("❌ Failed to Send")
        failed_df = pd.DataFrame(results['failed'])
        st.dataframe(failed_df, use_container_width=True)
    
    # Download updated file
    st.markdown("---")
    if st.button("📥 Download Updated Excel File"):
        with open(st.session_state.uploaded_file_path, 'rb') as f:
            st.download_button(
                label="💾 Download",
                data=f,
                file_name=f"interviews_updated_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    # Clear results
    if st.button("🔄 Start New Session"):
        st.session_state.email_results = []
        st.session_state.show_results = False
        st.session_state.uploaded_file_path = None
        st.rerun()

# Footer
st.markdown("---")
st.markdown("""
    <div style='text-align: center; color: #666; padding: 1rem;'>
        <p>💡 Make sure Outlook is installed and configured before sending emails</p>
        <p>📝 All activities are logged in <code>streamlit_email_notifications.log</code></p>
    </div>
""", unsafe_allow_html=True)
