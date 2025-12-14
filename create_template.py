"""
Excel Template Creation Script
Run this once to create the template Excel file with proper headers
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime, timedelta

def create_template():
    """Create a template Excel file with sample data"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Interviews"
    
    # Define headers
    headers = ["Candidate Email", "Interview Date", "Interview Time", "Interview Description", "Status"]
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Add headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    
    # Add sample data
    sample_data = [
        ["candidate1@example.com", (datetime.now() + timedelta(days=3)).strftime("%Y-%m-%d"), 
         "10:00 AM", "Technical Interview - Python & System Design", ""],
        ["candidate2@example.com", (datetime.now() + timedelta(days=5)).strftime("%Y-%m-%d"), 
         "2:00 PM", "HR Round - Cultural Fit Discussion", ""],
        ["candidate3@example.com", (datetime.now() + timedelta(days=7)).strftime("%Y-%m-%d"), 
         "11:30 AM", "Final Round - Meet the Team", ""],
    ]
    
    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 15
    
    # Save the file
    wb.save("template_interviews.xlsx")
    print("✓ Template file 'template_interviews.xlsx' created successfully!")
    print("  Please update with real candidate emails before running the main script.")

if __name__ == "__main__":
    create_template()
