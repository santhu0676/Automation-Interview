"""
Excel Reader Module
Handles reading interview data from Excel file
"""
import openpyxl
from typing import List, Dict
import os


class ExcelReader:
    """Reads and manages interview data from Excel file"""
    
    def __init__(self, file_path: str):
        """
        Initialize Excel reader
        
        Args:
            file_path: Path to the Excel file
        """
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
        
    def load_file(self) -> bool:
        """
        Load the Excel file
        
        Returns:
            True if file loaded successfully, False otherwise
        """
        try:
            if not os.path.exists(self.file_path):
                print(f"✗ Error: File '{self.file_path}' not found!")
                return False
                
            self.workbook = openpyxl.load_workbook(self.file_path)
            self.worksheet = self.workbook.active
            print(f"✓ Excel file loaded: {self.file_path}")
            return True
        except Exception as e:
            print(f"✗ Error loading Excel file: {str(e)}")
            return False
    
    def get_pending_interviews(self) -> List[Dict]:
        """
        Get all interviews that haven't been sent yet
        
        Returns:
            List of dictionaries containing interview details
        """
        interviews = []
        
        if not self.worksheet:
            return interviews
        
        # Skip header row (row 1)
        for row_num in range(2, self.worksheet.max_row + 1):
            # Get values from cells
            email = self.worksheet.cell(row=row_num, column=1).value
            date = self.worksheet.cell(row=row_num, column=2).value
            time = self.worksheet.cell(row=row_num, column=3).value
            description = self.worksheet.cell(row=row_num, column=4).value
            status = self.worksheet.cell(row=row_num, column=5).value
            
            # Skip if already sent or if email is empty
            if status == "Sent" or not email:
                continue
            
            # Validate required fields
            if not all([email, date, time, description]):
                print(f"⚠ Warning: Row {row_num} has missing data, skipping...")
                continue
            
            interviews.append({
                'row_num': row_num,
                'email': str(email).strip(),
                'date': str(date).strip(),
                'time': str(time).strip(),
                'description': str(description).strip()
            })
        
        return interviews
    
    def mark_as_sent(self, row_num: int) -> bool:
        """
        Mark an interview as sent in the Excel file
        
        Args:
            row_num: Row number to update
            
        Returns:
            True if updated successfully, False otherwise
        """
        try:
            self.worksheet.cell(row=row_num, column=5, value="Sent")
            self.workbook.save(self.file_path)
            return True
        except Exception as e:
            print(f"✗ Error marking row {row_num} as sent: {str(e)}")
            return False
    
    def close(self):
        """Close the workbook"""
        if self.workbook:
            self.workbook.close()
